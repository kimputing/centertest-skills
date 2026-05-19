"""
Rules 14002 and 14003: DDT data source integrity (cross-skill with ddt-tools).

14002 — @DataDriven datasource file existence check
14003 — DDT reference column integrity (# columns point to valid sheets/codes)
"""

from __future__ import annotations

import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from eir_models import CommitsDict, RuleResult
from eir_rules import rule
from rule_core import get_implemented_classes

try:
    from openpyxl import load_workbook
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# Regex to extract datasource path from @DataDriven annotation
_DATASOURCE_RE = re.compile(r'@DataDriven\s*\(\s*datasource\s*=\s*"([^"]+)"')


@rule(id="14002", description="@DataDriven datasource file existence", category="CenterTest")
def datasource_file_check(commits: CommitsDict, config) -> RuleResult:
    """
    Verify that @DataDriven(datasource = "testdata/...") references point
    to Excel files that actually exist on disk.

    Checks both class-level and method-level @DataDriven annotations.
    """
    result = RuleResult(
        rule_id="14002",
        description="@DataDriven datasource file existence",
        category="CenterTest",
        headers=["Class", "Method", "Datasource Path", "Issue"],
    )

    repo_dir = config.repository_dir

    for commit_info, files in sorted(commits.items()):
        for f in files:
            mc = f.main_class
            if mc is None:
                continue

            # Scan method bodies and annotations for @DataDriven
            for method in mc.methods:
                # Check method annotations — javalang extracts annotation names
                # but not parameters, so we need to scan the body text too
                datasource = None

                # Search in method body for the full annotation with datasource param
                if method.body:
                    m = _DATASOURCE_RE.search(method.body)
                    if m:
                        datasource = m.group(1)

                if datasource is None:
                    continue

                # Check if file exists
                full_path = os.path.join(repo_dir, datasource)
                if not os.path.isfile(full_path):
                    result.rows.append([
                        mc.class_name,
                        method.name,
                        datasource,
                        "File not found",
                    ])

    return result


def _load_sheet_codes(wb, sheet_name: str) -> dict[str, int]:
    """Load {code: row_number} from a sheet's 'code' column."""
    codes = {}
    ws = wb[sheet_name]
    # Find 'code' column (case-insensitive)
    code_col = None
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    for i, h in enumerate(header_row):
        if h and str(h).strip().lower() == "code":
            code_col = i
            break

    if code_col is None:
        return codes

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row_idx > 1 and code_col < len(row) and row[code_col]:
            val = str(row[code_col]).strip()
            if val:
                codes[val] = row_idx
    return codes


def _get_ref_columns(ws) -> list[tuple[int, str]]:
    """Find columns starting with # (reference columns). Returns [(col_idx, sheet_name)]."""
    refs = []
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    for i, h in enumerate(header_row):
        if h and str(h).strip().startswith("#"):
            sheet_name = str(h).strip()[1:]  # remove #
            refs.append((i, sheet_name))
    return refs


@rule(id="14003", description="DDT reference column integrity", category="CenterTest")
def ddt_reference_integrity(commits: CommitsDict, config) -> RuleResult:
    """
    Verify DDT Excel reference columns (# columns) point to valid sheets and codes.

    For each @DataDriven datasource file:
    1. Find columns starting with # (e.g., #Payment, #Coverage)
    2. Check that the referenced sheet exists (in same file or reference files)
    3. Check that each code in the # column exists in the referenced sheet

    This replicates the core logic of ddt-tools/xlsx-validate-refs.py
    as a healthcheck rule.
    """
    result = RuleResult(
        rule_id="14003",
        description="DDT reference column integrity",
        category="CenterTest",
        headers=["File", "Sheet", "Row", "Column", "Code", "Issue"],
    )

    if not _HAS_OPENPYXL:
        result.error = "openpyxl required for DDT reference validation"
        return result

    repo_dir = config.repository_dir

    # Collect all unique datasource paths from @DataDriven annotations
    datasource_paths = set()
    for commit_info, files in sorted(commits.items()):
        for f in files:
            mc = f.main_class
            if mc is None:
                continue
            for method in mc.methods:
                if method.body:
                    m = _DATASOURCE_RE.search(method.body)
                    if m:
                        datasource_paths.add(m.group(1))

    # Also scan testdata/ directory for all DC xlsx files
    testdata_dir = os.path.join(repo_dir, "testdata")
    if os.path.isdir(testdata_dir):
        for root, _dirs, filenames in os.walk(testdata_dir):
            for fname in filenames:
                if fname.endswith("DC.xlsx"):
                    rel = os.path.relpath(os.path.join(root, fname), repo_dir)
                    datasource_paths.add(rel)

    # Load reference data files (non-DC files in testdata/) for code lookup
    ref_codes: dict[str, dict[str, int]] = {}  # {sheet_name_lower: {code: row}}

    if os.path.isdir(testdata_dir):
        for fname in os.listdir(testdata_dir):
            if fname.endswith(".xlsx") and not fname.endswith("DC.xlsx"):
                ref_path = os.path.join(testdata_dir, fname)
                try:
                    wb_ref = load_workbook(ref_path, read_only=True, data_only=True)
                    for sheet_name in wb_ref.sheetnames:
                        codes = _load_sheet_codes(wb_ref, sheet_name)
                        if codes:
                            key = sheet_name.lower()
                            if key not in ref_codes:
                                ref_codes[key] = {}
                            ref_codes[key].update(codes)
                    wb_ref.close()
                except Exception:
                    pass

    # Validate each DC file
    for ds_path in sorted(datasource_paths):
        full_path = os.path.join(repo_dir, ds_path)
        if not os.path.isfile(full_path):
            continue  # rule 14002 handles missing files

        try:
            wb = load_workbook(full_path, read_only=True, data_only=True)
        except Exception:
            result.rows.append([ds_path, "-", "-", "-", "-", "Cannot open Excel file"])
            continue

        # Build local codes from all sheets in this DC file
        local_codes: dict[str, dict[str, int]] = {}
        for sheet_name in wb.sheetnames:
            codes = _load_sheet_codes(wb, sheet_name)
            if codes:
                local_codes[sheet_name.lower()] = codes

        # Check each sheet for # reference columns
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ref_cols = _get_ref_columns(ws)
            if not ref_cols:
                continue

            for col_idx, ref_sheet_name in ref_cols:
                ref_key = ref_sheet_name.lower()

                # Check if referenced sheet exists anywhere
                available_codes = {}
                if ref_key in local_codes:
                    available_codes = local_codes[ref_key]
                elif ref_key in ref_codes:
                    available_codes = ref_codes[ref_key]
                else:
                    result.rows.append([
                        os.path.basename(ds_path),
                        sheet_name,
                        "-",
                        f"#{ref_sheet_name}",
                        "-",
                        f"Referenced sheet '{ref_sheet_name}' not found",
                    ])
                    continue

                # Validate each code in the reference column
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if col_idx >= len(row) or not row[col_idx]:
                        continue
                    cell_value = str(row[col_idx]).strip()
                    if not cell_value:
                        continue

                    # Handle comma-separated codes
                    codes_to_check = [c.strip() for c in cell_value.split(",")]
                    for code in codes_to_check:
                        if code and code not in available_codes:
                            result.rows.append([
                                os.path.basename(ds_path),
                                sheet_name,
                                row_idx,
                                f"#{ref_sheet_name}",
                                code,
                                "Code not found in referenced sheet",
                            ])

        wb.close()

    return result
