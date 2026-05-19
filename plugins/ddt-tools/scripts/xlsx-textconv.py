#!/usr/bin/env python3
"""Convert xlsx files to readable text for git diff (textconv driver).

Usage:
    xlsx-textconv.py <file.xlsx>

Latest version: https://github.com/Kimputing/centertest-skills/blob/main/skills/ddt-tools/scripts/xlsx-textconv.py
"""

import sys
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def format_value(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)


def convert(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in sorted(wb.sheetnames):
        ws = wb[sheet_name]
        print(f"=== Sheet: {sheet_name} ===")
        headers = None
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [format_value(c) for c in row]
            if all(v == "" for v in values):
                continue
            if headers is None:
                headers = values
                print(f"[Header] {' | '.join(headers)}")
            else:
                pairs = [f"{h}={v}" for h, v in zip(headers, values) if v != ""]
                if pairs:
                    print(f"[Row {row_idx}] {' | '.join(pairs)}")
        print()
    wb.close()


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f"Usage: {sys.argv[0]} <file.xlsx>", file=sys.stderr)
        sys.exit(1)
    convert(sys.argv[1])
