#!/usr/bin/env python3
"""Validate that hardcoded DDTHelper string codes in Java source reference existing codes in xlsx files.

Parses DDTHelper.java to build a mapping of method -> (xlsx file, sheet), then scans
all Java source files for DDTHelper.getXxx("literal") calls and checks if the literal
code exists in the corresponding xlsx file/sheet.

Usage:
    ddt-check-code-usages.py                    # check all Java source files
    ddt-check-code-usages.py src/main/java/com/copperpoint/tests/SomeTest.java  # check specific file

Latest version: https://github.com/Kimputing/centertest-skills/blob/main/skills/ddt-tools/scripts/ddt-check-code-usages.py
"""

import glob
import os
import re
import sys
from datetime import datetime

# Add script directory to path for config import
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ddt_config import get_project_dir

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def format_value(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)


def parse_ddthelper(ddthelper_path):
    """Parse DDTHelper.java to extract method -> (xlsx_file, sheet_or_none) mapping.

    Two patterns:
      1. DC methods:  getExcelInputFileByCode("testdata/X.xlsx", code)  -> sheet=None (DataCombination)
      2. Ref methods: provideDataRecordAsExcelInput("testdata/X.xlsx", "sheet", code) -> sheet="sheet"

    Only maps methods with String code parameter (not ScenarioContext).
    """
    method_map = {}

    with open(ddthelper_path) as f:
        content = f.read()

    # Pattern 1: DC methods - getExcelInputFileByCode("file.xlsx", code)
    dc_pattern = re.compile(
        r'public static \S+ (\w+)\(String code\).*?'
        r'getExcelInputFileByCode\("([^"]+)",\s*code\)'
    )
    for m in dc_pattern.finditer(content):
        method_name = m.group(1)
        xlsx_file = m.group(2)
        method_map[method_name] = (xlsx_file, None)  # None = DataCombination sheet

    # Pattern 2: Reference methods - provideDataRecordAsExcelInput("file.xlsx", "sheet", code)
    ref_pattern = re.compile(
        r'public static \S+ (\w+)\(String code\).*?'
        r'provideDataRecordAsExcelInput\("([^"]+)",\s*"([^"]+)",\s*code\)'
    )
    for m in ref_pattern.finditer(content):
        method_name = m.group(1)
        xlsx_file = m.group(2)
        sheet_name = m.group(3)
        method_map[method_name] = (xlsx_file, sheet_name)

    return method_map


def load_codes_from_xlsx(xlsx_path, sheet_name=None):
    """Load all Code values from an xlsx file/sheet.

    If sheet_name is None, reads the DataCombination sheet.
    Returns dict of {lowercase_code: original_code} or None if file not found.
    """
    if not os.path.isfile(xlsx_path):
        return None  # File not found

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return None

    # Find the target sheet
    target = None
    if sheet_name is None:
        # Look for DataCombination sheet
        for name in wb.sheetnames:
            if name.lower() == "datacombination":
                target = name
                break
    else:
        # Look for sheet by name (case-insensitive)
        for name in wb.sheetnames:
            if name.lower() == sheet_name.lower():
                target = name
                break

    if target is None:
        wb.close()
        return {}

    ws = wb[target]
    headers = None
    codes = {}
    for row in ws.iter_rows(values_only=True):
        values = [format_value(c) for c in row]
        if all(v == "" for v in values):
            continue
        if headers is None:
            headers = [v.lower() for v in values]
            continue
        row_dict = dict(zip(headers, values))
        code = row_dict.get("code", "")
        if code:
            codes[code.lower()] = code

    wb.close()
    return codes


def find_ddthelper_calls(java_files):
    """Scan Java files for DDTHelper.method("literal") calls.

    Returns list of (java_file, line_num, method_name, code_literal).
    """
    # Match DDTHelper.methodName("string literal")
    pattern = re.compile(r'DDTHelper\.(\w+)\(\s*"([^"]+)"\s*\)')
    calls = []

    for java_file in java_files:
        try:
            with open(java_file) as f:
                for line_num, line in enumerate(f, start=1):
                    for m in pattern.finditer(line):
                        method_name = m.group(1)
                        code_literal = m.group(2)
                        calls.append((java_file, line_num, method_name, code_literal))
        except Exception:
            continue

    return calls


def main():
    project_dir = get_project_dir()
    os.chdir(project_dir)

    # Find DDTHelper.java
    ddthelper_candidates = glob.glob("src/**/DDTHelper.java", recursive=True)
    if not ddthelper_candidates:
        print("ERROR: DDTHelper.java not found in src/")
        sys.exit(1)

    ddthelper_path = ddthelper_candidates[0]
    print(f"Parsing {ddthelper_path}...")
    method_map = parse_ddthelper(ddthelper_path)
    print(f"  Found {len(method_map)} DDTHelper methods with String code parameter\n")

    # Find Java files to scan
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if args:
        java_files = args
    else:
        java_files = glob.glob("src/**/*.java", recursive=True)
        # Exclude DDTHelper itself and generated files
        java_files = [f for f in java_files if "DDTHelper.java" not in f]

    print(f"Scanning {len(java_files)} Java files for DDTHelper calls...")
    calls = find_ddthelper_calls(java_files)
    print(f"  Found {len(calls)} hardcoded DDTHelper.getXxx(\"code\") calls\n")

    if not calls:
        print("No hardcoded DDTHelper string code calls found.")
        return

    # Cache xlsx codes to avoid re-reading
    # Key: (xlsx_file, sheet_name) -> set of codes
    codes_cache = {}

    errors = []
    checked = 0

    for java_file, line_num, method_name, code_literal in calls:
        if method_name not in method_map:
            errors.append((java_file, line_num, method_name, code_literal,
                           "unknown DDTHelper method (not in DDTHelper.java)", []))
            continue

        xlsx_file, sheet_name = method_map[method_name]
        cache_key = (xlsx_file, sheet_name)

        if cache_key not in codes_cache:
            codes_cache[cache_key] = load_codes_from_xlsx(xlsx_file, sheet_name)

        available = codes_cache[cache_key]

        if available is None:
            errors.append((java_file, line_num, method_name, code_literal,
                           f"xlsx file not found: {xlsx_file}", []))
            continue

        checked += 1

        if code_literal.lower() not in available:
            sheet_label = sheet_name or "DataCombination"
            available_list = sorted(available.values())
            errors.append((java_file, line_num, method_name, code_literal,
                           f"code not found in {xlsx_file} sheet '{sheet_label}'",
                           available_list))

    # Report
    if not errors:
        print(f"All {checked} hardcoded DDTHelper codes are valid.")
    else:
        print(f"Found {len(errors)} issue(s):\n")
        current_file = None
        for java_file, line_num, method_name, code_literal, message, available_codes in errors:
            if java_file != current_file:
                current_file = java_file
                print(f"  {java_file}")
            print(f"    [{line_num}] DDTHelper.{method_name}(\"{code_literal}\") — {message}")
            if available_codes:
                print(f"           Available codes: {', '.join(available_codes)}")
        print(f"\n  {len(errors)} issue(s) in {len(set(e[0] for e in errors))} file(s).")
        print(f"  {checked} code(s) checked successfully.")
        sys.exit(1)


if __name__ == "__main__":
    main()
