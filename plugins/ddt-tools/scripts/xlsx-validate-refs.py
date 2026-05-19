#!/usr/bin/env python3
"""Validate that DC DataCombination sheets only reference codes that exist in their reference data sheets.

Usage:
    xlsx-validate-refs.py                          # validate all DC files
    xlsx-validate-refs.py testdata/WorkersCompDC.xlsx  # validate specific DC file

Latest version: https://github.com/Kimputing/centertest-skills/blob/main/skills/ddt-tools/scripts/xlsx-validate-refs.py
"""

import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

# Add script directory to path for config import
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ddt_config import get_project_dir

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def format_value(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)


def read_sheet(wb, sheet_name):
    """Read a sheet into {code: row_idx} mapping. Returns (headers, codes_dict)."""
    ws = wb[sheet_name]
    headers = None
    codes = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [format_value(c) for c in row]
        if all(v == "" for v in values):
            continue
        if headers is None:
            headers = [v.lower() for v in values]
            continue
        row_dict = dict(zip(headers, values))
        code = row_dict.get("code", "")
        if code:
            codes[code] = row_idx
    return headers, codes


def load_reference_codes(dc_path):
    """Load all available codes from all reference data files for a DC file.

    Returns: {sheet_name_lower: {code: (source_file, row_idx)}}
    """
    wb = load_workbook(dc_path, read_only=True, data_only=True)

    # Collect codes from sheets within the DC file itself (excluding DataCombination and References)
    all_codes = {}
    for sheet_name in wb.sheetnames:
        lower = sheet_name.lower()
        if lower in ("datacombination", "references"):
            continue
        _, codes = read_sheet(wb, sheet_name)
        if lower not in all_codes:
            all_codes[lower] = {}
        for code, row_idx in codes.items():
            all_codes[lower][code] = (os.path.basename(dc_path), row_idx)

    # Read References sheet to find external data files
    if "References" in wb.sheetnames:
        ref_headers, ref_codes = read_sheet(wb, "References")
        ws = wb["References"]
        headers = None
        for row in ws.iter_rows(values_only=True):
            values = [format_value(c) for c in row]
            if all(v == "" for v in values):
                continue
            if headers is None:
                headers = [v.lower() for v in values]
                continue
            row_dict = dict(zip(headers, values))
            location = row_dict.get("location", "")
            if location:
                ref_path = os.path.normpath(os.path.join(os.path.dirname(dc_path), location))
                if not os.path.isfile(ref_path):
                    ref_path = location
                if os.path.isfile(ref_path):
                    ref_wb = load_workbook(ref_path, read_only=True, data_only=True)
                    for sheet_name in ref_wb.sheetnames:
                        lower = sheet_name.lower()
                        _, codes = read_sheet(ref_wb, sheet_name)
                        if lower not in all_codes:
                            all_codes[lower] = {}
                        for code, row_idx in codes.items():
                            if code not in all_codes[lower]:
                                all_codes[lower][code] = (os.path.basename(ref_path), row_idx)
                    ref_wb.close()

    wb.close()

    # Also load from parent DC if this is a child DC (via hierarchy)
    hierarchy = load_hierarchy()
    for parent_dc, child_patterns in hierarchy.items():
        for pattern in child_patterns:
            if re.match(pattern, dc_path):
                if os.path.isfile(parent_dc):
                    parent_codes = load_reference_codes(parent_dc)
                    for sheet_lower, codes in parent_codes.items():
                        if sheet_lower not in all_codes:
                            all_codes[sheet_lower] = {}
                        for code, source in codes.items():
                            if code not in all_codes[sheet_lower]:
                                all_codes[sheet_lower][code] = source
                break

    return all_codes


_hierarchy_cache = None


def load_hierarchy():
    global _hierarchy_cache
    if _hierarchy_cache is not None:
        return _hierarchy_cache
    hierarchy_path = "testdata/DataDrivenHierarchy.json"
    if os.path.isfile(hierarchy_path):
        with open(hierarchy_path) as f:
            _hierarchy_cache = json.load(f)
    else:
        _hierarchy_cache = {}
    return _hierarchy_cache


def validate_dc(dc_path):
    """Validate a single DC file. Returns list of (row, col, code, sheet, message) errors."""
    if not os.path.isfile(dc_path):
        return [(0, "", "", "", f"File not found: {dc_path}")]

    try:
        wb = load_workbook(dc_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  WARNING: Skipping {dc_path} — cannot open: {e}", file=sys.stderr)
        return []

    # Find DataCombination sheet (case-insensitive)
    dc_sheet = None
    for name in wb.sheetnames:
        if name.lower() == "datacombination":
            dc_sheet = name
            break

    if dc_sheet is None:
        wb.close()
        return []

    # Read DataCombination headers
    ws = wb[dc_sheet]
    headers = None
    dc_rows = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [format_value(c) for c in row]
        if all(v == "" for v in values):
            continue
        if headers is None:
            headers = values
        else:
            dc_rows.append((row_idx, dict(zip(headers, values))))

    wb.close()

    if not headers:
        return []

    # Find reference columns (start with #)
    ref_columns = [h for h in headers if h.startswith("#")]
    if not ref_columns:
        return []

    # Load all available codes from reference data
    ref_codes = load_reference_codes(dc_path)

    errors = []

    # First pass: check which # columns have no matching sheet at all
    missing_sheets = set()
    for col in ref_columns:
        sheet_name = col[1:].lower()
        if sheet_name not in ref_codes:
            has_values = any(row_data.get(col, "") for _, row_data in dc_rows)
            if has_values:
                missing_sheets.add(col)
                errors.append((0, col, "", sheet_name, "no matching sheet found in any reference file"))

    # Second pass: validate codes for columns that do have a matching sheet
    for row_idx, row_data in dc_rows:
        for col in ref_columns:
            if col in missing_sheets:
                continue
            cell_value = row_data.get(col, "")
            if not cell_value:
                continue

            sheet_name = col[1:].lower()
            available = ref_codes[sheet_name]

            codes = [c.strip() for c in cell_value.split(",")]
            for code in codes:
                if code and code not in available:
                    similar = [c for c in available if c.lower() == code.lower()]
                    hint = f" (did you mean '{similar[0]}'?)" if similar else ""
                    errors.append((row_idx, col, code, sheet_name, f"code not found{hint}"))

    return errors


def find_all_dc_files():
    """Find all DC xlsx files in testdata/."""
    import glob
    files = glob.glob("testdata/**/*DC.xlsx", recursive=True) + glob.glob("testdata/*DC.xlsx")
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    return sorted(set(files))


def main():
    project_dir = get_project_dir()
    os.chdir(project_dir)

    args = [a for a in sys.argv[1:] if not a.startswith("--")]

    if args:
        dc_files = args
    else:
        dc_files = find_all_dc_files()

    total_errors = 0
    files_with_errors = 0

    for dc_path in dc_files:
        try:
            errors = validate_dc(dc_path)
        except Exception as e:
            print(f"\n  {dc_path}")
            print(f"    ERROR: failed to validate — {e}")
            files_with_errors += 1
            total_errors += 1
            continue
        if errors:
            files_with_errors += 1
            print(f"\n  {dc_path}")
            for row_idx, col, code, sheet, message in errors:
                total_errors += 1
                print(f"    [Row {row_idx}] {col} -> '{code}' : {message} (sheet: {sheet})")

    if total_errors == 0:
        print("  All DC references are valid.")
    else:
        print(f"\n  Found {total_errors} broken reference(s) in {files_with_errors} file(s).")
        sys.exit(1)


if __name__ == "__main__":
    main()
