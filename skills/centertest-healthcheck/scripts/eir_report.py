#!/usr/bin/env python3
"""
Report generator for eir-analyzer.

Produces Excel (.xlsx), Markdown (.md), and terminal output.
Follows the ddt-analyzer pattern for Excel generation with openpyxl.

Latest version: https://github.com/Kimputing/centertest-skills/blob/main/skills/eir-analyzer/scripts/eir_report.py
"""

from __future__ import annotations

import os
import sys
from datetime import datetime

from eir_models import RuleResult, Section

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None


# ---------------------------------------------------------------------------
# Sheet name safety
# ---------------------------------------------------------------------------

_FORBIDDEN_CHARS = set("[]:\\/*?")


def _safe_sheet_name(rule_id: str, description: str, existing_names: set[str]) -> str:
    """Generate a unique, valid Excel sheet name (max 31 chars)."""
    short_desc = description[:22].strip()
    base = f"R{rule_id}_{short_desc}"
    # Remove forbidden characters
    base = "".join(c for c in base if c not in _FORBIDDEN_CHARS)
    base = base[:31]

    name = base
    counter = 2
    while name in existing_names:
        suffix = f"_{counter}"
        name = base[:31 - len(suffix)] + suffix
        counter += 1
    return name


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def generate_excel(results: list[RuleResult], output_dir: str) -> str:
    """Generate an Excel report with one sheet per rule + summary sheet."""
    if Workbook is None:
        print("  Warning: openpyxl not installed, skipping Excel output.", file=sys.stderr)
        return ""

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    error_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_headers = ["Rule ID", "Description", "Category", "Items Found", "Status"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    existing_names = {"Summary"}

    for i, result in enumerate(results):
        # Summary row
        if result.error:
            items = "ERROR"
            status = result.error[:50]
        elif result.rows:
            items = len(result.rows)
            status = "Issues found" if items > 0 else "OK"
        elif result.sections:
            items = sum(len(s.items) + sum(len(ss.items) for ss in s.subsections) for s in result.sections)
            status = "Issues found" if items > 0 else "OK"
        else:
            items = 0
            status = "OK"

        row_num = i + 2
        ws_summary.cell(row=row_num, column=1, value=result.rule_id)
        ws_summary.cell(row=row_num, column=2, value=result.description)
        ws_summary.cell(row=row_num, column=3, value=result.category)
        ws_summary.cell(row=row_num, column=4, value=items)
        cell = ws_summary.cell(row=row_num, column=5, value=status)
        if result.error:
            cell.fill = error_fill

        # Skip detail sheet if no content (no rows, no sections, no error)
        has_content = result.error or result.rows or result.sections
        if not has_content:
            continue

        # Detail sheet
        sheet_name = _safe_sheet_name(result.rule_id, result.description, existing_names)
        existing_names.add(sheet_name)
        ws = wb.create_sheet(title=sheet_name)

        if result.error:
            ws.cell(row=1, column=1, value=f"Rule {result.rule_id} failed:")
            ws.cell(row=2, column=1, value=result.error)
            continue

        next_row = 1
        if result.headers and result.rows:
            # Flat tabular output
            for col, header in enumerate(result.headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            for row_idx, row_data in enumerate(result.rows, 2):
                for col, val in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col, value=val)
            next_row = len(result.rows) + 3  # blank row after table

        if result.sections:
            # Hierarchical output — append after table (or write standalone)
            _write_sections_to_sheet(ws, result.sections, header_font, header_fill,
                                     start_row=next_row)

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 120)

    # Auto-width summary columns
    for col in ws_summary.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_summary.column_dimensions[col_letter].width = min(max_len + 2, 120)

    # Save
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"HealthCheck_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


def _write_sections_to_sheet(ws, sections: list[Section], header_font, header_fill,
                             start_row: int = 1):
    """Write hierarchical sections to an Excel sheet."""
    row = start_row
    section_font = Font(bold=True, size=12)
    subsection_font = Font(bold=True, size=11)

    for section in sections:
        # Section title
        cell = ws.cell(row=row, column=1, value=section.title)
        cell.font = section_font
        row += 1

        # Section table (if present)
        if section.headers and section.rows:
            for col, header in enumerate(section.headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            row += 1
            for row_data in section.rows:
                for col, val in enumerate(row_data, 1):
                    ws.cell(row=row, column=col, value=val)
                row += 1

        # Section items
        for item in section.items:
            ws.cell(row=row, column=2, value=item)
            row += 1

        # Subsections
        for sub in section.subsections:
            cell = ws.cell(row=row, column=2, value=sub.title)
            cell.font = subsection_font
            row += 1
            for item in sub.items:
                ws.cell(row=row, column=3, value=item)
                row += 1

        row += 1  # blank row between sections


# ---------------------------------------------------------------------------
# Markdown report
# ---------------------------------------------------------------------------

def generate_markdown(results: list[RuleResult], output_dir: str) -> str:
    """Generate a Markdown report."""
    lines = ["# CenterTest Health Check Report", ""]
    lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("")

    # Summary table
    lines.append("## Summary")
    lines.append("")
    lines.append("| Rule | Description | Items | Status |")
    lines.append("|------|-------------|-------|--------|")

    for result in results:
        if result.error:
            items = "ERROR"
            status = result.error[:40]
        elif result.rows:
            items = str(len(result.rows))
            status = "Issues" if result.rows else "OK"
        elif result.sections:
            items = str(sum(len(s.items) + sum(len(ss.items) for ss in s.subsections) for s in result.sections))
            status = "Issues" if int(items) > 0 else "OK"
        else:
            items = "0"
            status = "OK"
        lines.append(f"| {result.rule_id} | {result.description} | {items} | {status} |")

    lines.append("")

    # Detail sections
    for result in results:
        lines.append(f"## Rule {result.rule_id}: {result.description}")
        lines.append("")

        if result.error:
            lines.append(f"**ERROR:** {result.error}")
            lines.append("")
            continue

        if result.headers and result.rows:
            lines.append("| " + " | ".join(str(h) for h in result.headers) + " |")
            lines.append("| " + " | ".join("---" for _ in result.headers) + " |")
            for row in result.rows:
                lines.append("| " + " | ".join(str(v) for v in row) + " |")
            lines.append("")

        if result.sections:
            for section in result.sections:
                lines.append(f"### {section.title}")
                lines.append("")
                if section.headers and section.rows:
                    lines.append("| " + " | ".join(str(h) for h in section.headers) + " |")
                    lines.append("| " + " | ".join("---" for _ in section.headers) + " |")
                    for row in section.rows:
                        lines.append("| " + " | ".join(str(v) for v in row) + " |")
                    lines.append("")
                for item in section.items:
                    lines.append(f"- {item}")
                for sub in section.subsections:
                    lines.append(f"#### {sub.title}")
                    for item in sub.items:
                        lines.append(f"  - {item}")
                lines.append("")
        else:
            lines.append("No issues found.")
            lines.append("")

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"HealthCheck_{timestamp}.md"
    filepath = os.path.join(output_dir, filename)
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return filepath


# ---------------------------------------------------------------------------
# HTML report
# ---------------------------------------------------------------------------

def _count_items(result: RuleResult) -> int:
    """Count total items in a RuleResult."""
    if result.error:
        return 0
    section_count = sum(len(s.items) + sum(len(ss.items) for ss in s.subsections) for s in result.sections)
    # If rule has both rows (summary table) and sections (detail findings),
    # count only the sections — the rows are informational summaries, not issues
    if result.sections and section_count > 0:
        return section_count
    return len(result.rows)


def _html_escape(text: str) -> str:
    """Escape HTML special characters."""
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def generate_html(results: list[RuleResult], output_dir: str,
                  project_name: str = "", total_files: int = 0,
                  elapsed: float = 0, excel_filename: str = "") -> str:
    """Generate a dashboard-style HTML health check report."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Compute summary stats with weighted scoring
    # Rules have different severity weights:
    #   - Informational (Statistics): not scored — always have data, not issues
    #   - Low (ClassNames, Methods): suggestions, nice-to-have
    #   - Medium (Inheritance, Selenium, Complexity, CenterTest): should fix
    #   - High (Quality, Security): must fix
    _SEVERITY = {
        "Statistics": None,    # informational — excluded from scoring
        "ClassNames": "low",
        "Methods":    "low",
        "Inheritance": "medium",
        "Selenium":   "medium",
        "Complexity": "medium",
        "CenterTest": "medium",
        "Quality":    "high",
        "Security":   "high",
    }
    _WEIGHT = {"low": 1, "medium": 2, "high": 3}

    total_issues = 0
    rules_passed = 0
    rules_failed = 0
    rules_error = 0
    score_earned = 0.0
    score_possible = 0.0
    categories: dict[str, dict] = {}

    for r in results:
        count = _count_items(r)
        total_issues += count
        severity = _SEVERITY.get(r.category, "medium")

        if r.error:
            rules_error += 1
        elif count > 0:
            rules_failed += 1
        else:
            rules_passed += 1

        # Weighted scoring (skip informational rules)
        if severity is not None:
            weight = _WEIGHT[severity]
            score_possible += weight
            if count == 0 and not r.error:
                score_earned += weight
            elif not r.error:
                # Partial credit: few findings in a large codebase is still mostly healthy
                # Deduct proportionally but never below 0
                penalty = min(count / max(total_files, 1), 1.0)
                score_earned += weight * max(0, 1.0 - penalty)

        cat = r.category
        if cat not in categories:
            categories[cat] = {"passed": 0, "failed": 0, "issues": 0}
        if r.error or count > 0:
            categories[cat]["failed"] += 1
        else:
            categories[cat]["passed"] += 1
        categories[cat]["issues"] += count

    total_rules = len(results)
    health_score = round(score_earned / score_possible * 100) if score_possible > 0 else 0

    # Determine score color
    if health_score >= 80:
        score_color = "#22c55e"  # green
        score_label = "Healthy"
    elif health_score >= 60:
        score_color = "#f59e0b"  # amber
        score_label = "Needs Attention"
    else:
        score_color = "#ef4444"  # red
        score_label = "Critical"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CenterTest Health Check — {_html_escape(project_name)}</title>
<style>
  :root {{
    --bg: #0f172a; --surface: #1e293b; --surface2: #334155;
    --border: #475569; --text: #e2e8f0; --text-muted: #94a3b8;
    --accent: #3b82f6; --green: #22c55e; --amber: #f59e0b; --red: #ef4444;
  }}
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    background: var(--bg); color: var(--text); line-height: 1.6;
  }}
  .container {{ max-width: 1400px; margin: 0 auto; padding: 24px; }}

  /* Header */
  .header {{
    display: flex; justify-content: space-between; align-items: center;
    padding: 24px 0; border-bottom: 1px solid var(--border); margin-bottom: 32px;
  }}
  .header h1 {{ font-size: 24px; font-weight: 700; }}
  .header h1 span {{ color: var(--accent); }}
  .header-meta {{ text-align: right; color: var(--text-muted); font-size: 13px; }}

  /* Score card */
  .score-row {{
    display: grid; grid-template-columns: 200px 1fr; gap: 24px; margin-bottom: 32px;
  }}
  .score-card {{
    background: var(--surface); border-radius: 16px; padding: 32px;
    display: flex; flex-direction: column; align-items: center; justify-content: center;
    border: 2px solid {score_color};
  }}
  .score-number {{ font-size: 56px; font-weight: 800; color: {score_color}; line-height: 1; }}
  .score-label {{ font-size: 14px; color: var(--text-muted); margin-top: 4px; }}
  .score-badge {{
    margin-top: 12px; padding: 4px 16px; border-radius: 20px; font-size: 13px;
    font-weight: 600; background: {score_color}22; color: {score_color};
  }}

  /* Stat cards */
  .stats-grid {{
    display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
    gap: 16px;
  }}
  .stat-card {{
    background: var(--surface); border-radius: 12px; padding: 20px;
    border: 1px solid var(--border);
  }}
  .stat-value {{ font-size: 28px; font-weight: 700; }}
  .stat-label {{ font-size: 12px; color: var(--text-muted); text-transform: uppercase;
    letter-spacing: 0.05em; margin-top: 4px; }}
  .stat-green {{ color: var(--green); }}
  .stat-red {{ color: var(--red); }}
  .stat-amber {{ color: var(--amber); }}
  .stat-blue {{ color: var(--accent); }}

  /* Category bars */
  .categories {{ margin: 32px 0; }}
  .categories h2 {{ font-size: 18px; font-weight: 600; margin-bottom: 16px; }}
  .cat-row {{
    display: flex; align-items: center; gap: 12px; padding: 10px 0;
    border-bottom: 1px solid var(--surface2);
  }}
  .cat-name {{ width: 140px; font-weight: 500; font-size: 14px; }}
  .cat-bar-bg {{
    flex: 1; height: 24px; background: var(--surface2); border-radius: 12px;
    overflow: hidden; position: relative;
  }}
  .cat-bar-fill {{
    height: 100%; border-radius: 12px; transition: width 0.5s;
  }}
  .cat-bar-fill.green {{ background: var(--green); }}
  .cat-bar-fill.red {{ background: var(--red); }}
  .cat-bar-fill.amber {{ background: var(--amber); }}
  .cat-issues {{ width: 80px; text-align: right; font-size: 14px; color: var(--text-muted); }}

  /* Rules table */
  .rules-section {{ margin: 32px 0; }}
  .rules-section h2 {{ font-size: 18px; font-weight: 600; margin-bottom: 16px; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th {{
    text-align: left; padding: 12px 16px; font-size: 12px; text-transform: uppercase;
    letter-spacing: 0.05em; color: var(--text-muted); border-bottom: 2px solid var(--border);
    background: var(--surface);
  }}
  td {{ padding: 12px 16px; border-bottom: 1px solid var(--surface2); font-size: 14px; }}
  tr:hover td {{ background: var(--surface); }}
  .badge {{
    display: inline-block; padding: 2px 10px; border-radius: 10px;
    font-size: 12px; font-weight: 600;
  }}
  .badge-ok {{ background: var(--green)22; color: var(--green); }}
  .badge-issues {{ background: var(--red)22; color: var(--red); }}
  .badge-error {{ background: var(--amber)22; color: var(--amber); }}
  .badge-cat {{
    background: var(--accent)18; color: var(--accent); font-size: 11px;
    padding: 2px 8px; border-radius: 8px;
  }}

  /* Detail sections */
  .detail {{ margin: 32px 0; }}
  .detail h2 {{ font-size: 18px; font-weight: 600; margin-bottom: 16px; }}
  .rule-detail {{
    background: var(--surface); border-radius: 12px; border: 1px solid var(--border);
    margin-bottom: 16px; overflow: hidden;
  }}
  .rule-detail-header {{
    padding: 16px 20px; display: flex; justify-content: space-between;
    align-items: center; cursor: pointer; user-select: none;
  }}
  .rule-detail-header:hover {{ background: var(--surface2); }}
  .rule-detail-title {{ font-weight: 600; }}
  .rule-detail-body {{ padding: 0 20px 16px; }}
  .detail-table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
  .detail-table th {{
    text-align: left; padding: 8px 12px; font-size: 11px; background: var(--surface2);
    border-bottom: 1px solid var(--border);
  }}
  .detail-table td {{
    padding: 8px 12px; font-size: 13px; border-bottom: 1px solid var(--surface2);
    font-family: 'SF Mono', Monaco, monospace; word-break: break-all; max-width: 600px;
  }}
  .finding-list {{ list-style: none; padding: 0; }}
  .finding-list li {{
    padding: 6px 12px; font-size: 13px; border-left: 3px solid var(--accent);
    margin: 4px 0; background: var(--surface2); border-radius: 0 6px 6px 0;
    font-family: 'SF Mono', Monaco, monospace;
  }}
  .section-title {{
    font-weight: 600; font-size: 14px; margin: 12px 0 6px; color: var(--accent);
  }}
  .subsection-title {{
    font-weight: 500; font-size: 13px; margin: 8px 0 4px; color: var(--text-muted);
    padding-left: 12px;
  }}

  /* Footer */
  .footer {{
    margin-top: 48px; padding: 24px 0; border-top: 1px solid var(--border);
    color: var(--text-muted); font-size: 12px; text-align: center;
  }}
  .footer a {{ color: var(--accent); text-decoration: none; }}

  /* Toggle */
  .toggle-arrow {{ transition: transform 0.2s; font-size: 12px; color: var(--text-muted); }}
  .rule-detail.open .toggle-arrow {{ transform: rotate(90deg); }}
  .rule-detail:not(.open) .rule-detail-body {{ display: none; }}
</style>
</head>
<body>
<div class="container">

  <div class="header">
    <h1><span>CenterTest</span> Health Check</h1>
    <div class="header-meta">
      <div>{_html_escape(project_name)}</div>
      <div>{generated_at}</div>
    </div>
  </div>

  <div class="score-row">
    <div class="score-card">
      <div class="score-number">{health_score}</div>
      <div class="score-label">Health Score</div>
      <div class="score-badge">{score_label}</div>
    </div>
    <div class="stats-grid">
      <div class="stat-card">
        <div class="stat-value stat-blue">{total_rules}</div>
        <div class="stat-label">Rules Checked</div>
      </div>
      <div class="stat-card">
        <div class="stat-value stat-green">{rules_passed}</div>
        <div class="stat-label">Passed</div>
      </div>
      <div class="stat-card">
        <div class="stat-value stat-red">{rules_failed}</div>
        <div class="stat-label">Issues Found</div>
      </div>
      <div class="stat-card">
        <div class="stat-value stat-amber">{total_issues}</div>
        <div class="stat-label">Total Findings</div>
      </div>
      <div class="stat-card">
        <div class="stat-value stat-blue">{total_files}</div>
        <div class="stat-label">Files Analyzed</div>
      </div>
      <div class="stat-card">
        <div class="stat-value" style="color:var(--text-muted)">{elapsed:.1f}s</div>
        <div class="stat-label">Duration</div>
      </div>
    </div>
  </div>
"""

    # Category breakdown
    html += '  <div class="categories">\n    <h2>By Category</h2>\n'
    max_issues = max((c["issues"] for c in categories.values()), default=1) or 1
    for cat_name, cat_data in sorted(categories.items(), key=lambda x: -x[1]["issues"]):
        total_cat = cat_data["passed"] + cat_data["failed"]
        pct = cat_data["passed"] / total_cat * 100 if total_cat > 0 else 100
        bar_pct = cat_data["issues"] / max_issues * 100
        bar_class = "green" if cat_data["issues"] == 0 else ("amber" if cat_data["issues"] < 10 else "red")
        html += f"""    <div class="cat-row">
      <div class="cat-name">{_html_escape(cat_name)}</div>
      <div class="cat-bar-bg">
        <div class="cat-bar-fill {bar_class}" style="width:{bar_pct:.0f}%"></div>
      </div>
      <div class="cat-issues">{cat_data['issues']} issue{"s" if cat_data["issues"] != 1 else ""}</div>
    </div>
"""
    html += '  </div>\n'

    # Rules summary table (collapsible)
    html += """  <div class="rules-section">
    <div class="rule-detail open">
      <div class="rule-detail-header" onclick="this.parentElement.classList.toggle('open')">
        <span class="rule-detail-title" style="font-size:18px">Rules Summary</span>
        <span class="toggle-arrow">&#9654;</span>
      </div>
      <div class="rule-detail-body">
    <table>
      <thead><tr><th>Rule</th><th>Category</th><th>Description</th><th>Findings</th><th>Status</th></tr></thead>
      <tbody>
"""
    for r in results:
        count = _count_items(r)
        if r.error:
            badge = '<span class="badge badge-error">ERROR</span>'
        elif count > 0:
            badge = f'<span class="badge badge-issues">{count} issue{"s" if count != 1 else ""}</span>'
        else:
            badge = '<span class="badge badge-ok">Passed</span>'
        html += f"""        <tr>
          <td><strong>{r.rule_id}</strong></td>
          <td><span class="badge-cat">{_html_escape(r.category)}</span></td>
          <td>{_html_escape(r.description)}</td>
          <td>{count}</td>
          <td>{badge}</td>
        </tr>
"""
    html += "      </tbody>\n    </table>\n      </div>\n    </div>\n  </div>\n"

    # Detail sections (grouped by category, collapsible)
    html += '  <div class="detail">\n    <h2>Detailed Findings</h2>\n'

    # Group results by category, preserving order
    from collections import OrderedDict
    cat_results: OrderedDict[str, list] = OrderedDict()
    for r in results:
        cat_results.setdefault(r.category, []).append(r)

    for cat_name, cat_rules in cat_results.items():
        cat_issue_count = sum(_count_items(r) for r in cat_rules)
        cat_has_content = any(_count_items(r) > 0 or r.error for r in cat_rules)
        if not cat_has_content:
            continue

        # Category wrapper (collapsed by default)
        cat_open = ""
        html += f'    <div class="rule-detail{cat_open}">\n'
        html += f'      <div class="rule-detail-header" onclick="this.parentElement.classList.toggle(\'open\')">\n'
        html += f'        <span class="rule-detail-title" style="font-size:16px">{_html_escape(cat_name)}</span>\n'
        cat_rule_count = sum(1 for r in cat_rules if _count_items(r) > 0 or r.error)
        html += f'        <span><span style="margin-right:8px;color:var(--text-muted)">{cat_issue_count} finding{"s" if cat_issue_count != 1 else ""} in {cat_rule_count} rule{"s" if cat_rule_count != 1 else ""}</span><span class="toggle-arrow">&#9654;</span></span>\n'
        html += '      </div>\n'
        html += '      <div class="rule-detail-body">\n'

        # Individual rules within category
        for r in cat_rules:
            count = _count_items(r)
            if count == 0 and not r.error:
                continue

            open_class = ""  # collapsed by default
            html += f'        <div class="rule-detail{open_class}" style="margin-left:8px;border-color:var(--surface2)">\n'
            html += f'          <div class="rule-detail-header" onclick="this.parentElement.classList.toggle(\'open\')">\n'
            html += f'            <span class="rule-detail-title">{r.rule_id}: {_html_escape(r.description)}</span>\n'
            count_text = f"{count} finding{'s' if count != 1 else ''}" if not r.error else "ERROR"
            html += f'            <span><span style="margin-right:8px;color:var(--text-muted)">{count_text}</span><span class="toggle-arrow">&#9654;</span></span>\n'
            html += '          </div>\n'
            html += '          <div class="rule-detail-body">\n'

            if r.error:
                html += f'            <p style="color:var(--amber)">Error: {_html_escape(r.error)}</p>\n'

            # Flat table
            if r.headers and r.rows:
                html += '            <table class="detail-table"><thead><tr>'
                for h in r.headers:
                    html += f'<th>{_html_escape(h)}</th>'
                html += '</tr></thead><tbody>\n'
                for row in r.rows[:100]:
                    html += '              <tr>'
                    for val in row:
                        html += f'<td>{_html_escape(val)}</td>'
                    html += '</tr>\n'
                if len(r.rows) > 100:
                    html += f'              <tr><td colspan="{len(r.headers)}" style="color:var(--text-muted);text-align:center">... and {len(r.rows) - 100} more (see Excel)</td></tr>\n'
                html += '            </tbody></table>\n'

            # Sections
            for section in r.sections:
                html += f'            <div class="section-title">{_html_escape(section.title)}</div>\n'
                if section.headers and section.rows:
                    html += '            <table class="detail-table"><thead><tr>'
                    for h in section.headers:
                        html += f'<th>{_html_escape(h)}</th>'
                    html += '</tr></thead><tbody>\n'
                    for row in section.rows:
                        html += '              <tr>'
                        for val in row:
                            html += f'<td>{_html_escape(val)}</td>'
                        html += '</tr>\n'
                    html += '            </tbody></table>\n'
                if section.items:
                    html += '            <ul class="finding-list">\n'
                    for item in section.items[:50]:
                        html += f'              <li>{_html_escape(item)}</li>\n'
                    if len(section.items) > 50:
                        html += f'              <li style="color:var(--text-muted)">... and {len(section.items) - 50} more</li>\n'
                    html += '            </ul>\n'
                for sub in section.subsections:
                    html += f'            <div class="subsection-title">{_html_escape(sub.title)}</div>\n'
                    if sub.items:
                        html += '            <ul class="finding-list">\n'
                        for item in sub.items[:30]:
                            html += f'              <li>{_html_escape(item)}</li>\n'
                        if len(sub.items) > 30:
                            html += f'              <li style="color:var(--text-muted)">... and {len(sub.items) - 30} more</li>\n'
                        html += '            </ul>\n'

            html += '          </div>\n        </div>\n'  # close rule

        html += '      </div>\n    </div>\n'  # close category

    html += '  </div>\n'

    # Footer
    excel_link = f'<a href="{_html_escape(excel_filename)}">{_html_escape(excel_filename)}</a>' if excel_filename else ""
    html += f"""
  <div class="footer">
    CenterTest Health Check &mdash; Generated {generated_at}
    {f"&nbsp;&bull;&nbsp; Full data: {excel_link}" if excel_link else ""}
    <br>Powered by <a href="https://github.com/Kimputing/centertest-skills">centertest-skills</a>
  </div>

</div>
</body>
</html>
"""

    os.makedirs(output_dir, exist_ok=True)
    filename = f"HealthCheck_{timestamp}.html"
    filepath = os.path.join(output_dir, filename)
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(html)
    return filepath


# ---------------------------------------------------------------------------
# Terminal output
# ---------------------------------------------------------------------------

def print_terminal_summary(results: list[RuleResult]):
    """Print a concise summary to the terminal."""
    print("\n" + "=" * 60)
    print("  HEALTH CHECK RESULTS")
    print("=" * 60)

    total_issues = 0
    errors = 0

    for result in results:
        if result.error:
            errors += 1
            status = "ERROR"
            count = "-"
        elif result.rows:
            count = len(result.rows)
            total_issues += count
            status = f"{count} issue(s)" if count > 0 else "OK"
        elif result.sections:
            count = sum(len(s.items) + sum(len(ss.items) for ss in s.subsections) for s in result.sections)
            total_issues += count
            status = f"{count} finding(s)" if count > 0 else "OK"
        else:
            count = 0
            status = "OK"

        indicator = "X" if (result.error or (isinstance(count, int) and count > 0)) else "."
        print(f"  [{indicator}] {result.rule_id}: {result.description:<50} {status}")

    print("-" * 60)
    print(f"  Total issues: {total_issues}  |  Errors: {errors}  |  Rules: {len(results)}")
    print("=" * 60)
