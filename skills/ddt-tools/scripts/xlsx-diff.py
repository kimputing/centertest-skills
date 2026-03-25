#!/usr/bin/env python3
"""Smart cell-level diff for xlsx files against origin/main.

Usage:
    xlsx-diff.py <file.xlsx>                    # diff working copy vs origin/main
    xlsx-diff.py --ref <ref> <file.xlsx>        # diff working copy vs specific ref
    xlsx-diff.py <old.xlsx> <new.xlsx>          # diff two files directly

Latest version: https://github.com/Kimputing/centertest-skills/blob/main/skills/ddt-tools/scripts/xlsx-diff.py
"""

import os
import subprocess
import sys
import tempfile
from collections import defaultdict, OrderedDict
from datetime import datetime

# Add script directory to path for config import
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ddt_config import get_project_dir

DEFAULT_REF = "origin/main"

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def format_value(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)


def parse_xlsx(path):
    """Parse xlsx into {sheet_name: {"headers": [...], "rows": OrderedDict(code -> {header: value})}}."""
    result = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None
        rows = OrderedDict()
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [format_value(c) for c in row]
            if all(v == "" for v in values):
                continue
            if headers is None:
                headers = values
            else:
                row_dict = {}
                for h, v in zip(headers, values):
                    row_dict[h] = v
                code = row_dict.get("Code", "") or row_dict.get("code", "")
                key = code if code else f"__row_{row_idx}"
                if key in rows:
                    key = f"{key}__row_{row_idx}"
                rows[key] = row_dict
        result[sheet_name] = {"headers": headers or [], "rows": rows}
    wb.close()
    return result


def get_rel_path(filepath):
    """Get the repo-relative path for a file."""
    rel = subprocess.check_output(
        ["git", "ls-files", "--full-name", filepath], text=True
    ).strip()
    if rel:
        return rel
    repo_root = subprocess.check_output(
        ["git", "rev-parse", "--show-toplevel"], text=True
    ).strip()
    abs_path = os.path.abspath(filepath)
    if abs_path.startswith(repo_root):
        return abs_path[len(repo_root) + 1:]
    return filepath


def get_ref_version(filepath, ref=DEFAULT_REF):
    """Extract a git ref version of a file into a temp file."""
    rel_path = get_rel_path(filepath)
    try:
        content = subprocess.check_output(
            ["git", "show", f"{ref}:{rel_path}"],
        )
    except subprocess.CalledProcessError:
        return None
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(content)
    tmp.close()
    return tmp.name


def format_row_short(row_dict, headers):
    """Format a row as key=value pairs, non-empty only."""
    pairs = [f"{h}={row_dict.get(h, '')}" for h in headers if row_dict.get(h, '') != ""]
    return " | ".join(pairs)


def format_key_range(keys):
    """Format a list of code keys for display."""
    if len(keys) == 1:
        return keys[0]
    return ", ".join(keys)


def diff_sheets(old_data, new_data):
    """Compare two parsed xlsx structures and print cell-level diffs."""
    all_sheets = sorted(set(list(old_data.keys()) + list(new_data.keys())))
    has_changes = False

    for sheet in all_sheets:
        old_info = old_data.get(sheet, {"headers": [], "rows": OrderedDict()})
        new_info = new_data.get(sheet, {"headers": [], "rows": OrderedDict()})
        old_headers = old_info["headers"]
        new_headers = new_info["headers"]
        old_rows = old_info["rows"]
        new_rows = new_info["rows"]
        all_headers = new_headers or old_headers

        if sheet not in old_data:
            if not new_rows:
                continue
            has_changes = True
            codes = [r.get("Code", "") or r.get("code", "") for r in new_rows.values()]
            codes = [c for c in codes if c]
            print(f"  {sheet} (NEW SHEET) — {len(new_rows)} rows: {', '.join(codes)}")
            continue

        if sheet not in new_data:
            has_changes = True
            if old_rows:
                codes = [r.get("Code", "") or r.get("code", "") for r in old_rows.values()]
                codes = [c for c in codes if c]
                print(f"  {sheet} (REMOVED) — had {len(old_rows)} rows: {', '.join(codes)}")
            else:
                print(f"  {sheet} (REMOVED)")
            continue

        sheet_lines = []

        # Header changes
        if old_headers != new_headers:
            added_cols = [h for h in new_headers if h and h not in old_headers]
            removed_cols = [h for h in old_headers if h and h not in new_headers]
            if added_cols:
                sheet_lines.append(f"    columns added: {', '.join(added_cols)}")
            if removed_cols:
                sheet_lines.append(f"    columns removed: {', '.join(removed_cols)}")

        old_keys = set(old_rows.keys())
        new_keys = set(new_rows.keys())

        # Rows in both — check for cell changes
        change_groups = defaultdict(list)
        for key in old_keys & new_keys:
            old_row = old_rows[key]
            new_row = new_rows[key]
            all_keys_list = list(dict.fromkeys(list(all_headers)))
            cell_changes = []
            for col in all_keys_list:
                old_val = old_row.get(col, "")
                new_val = new_row.get(col, "")
                if old_val != new_val:
                    if old_val == "":
                        cell_changes.append(f"{col} added: '{new_val}'")
                    elif new_val == "":
                        cell_changes.append(f"{col} removed: '{old_val}'")
                    else:
                        cell_changes.append(f"{col}: '{old_val}' -> '{new_val}'")
            if cell_changes:
                sig = tuple(cell_changes)
                change_groups[sig].append(key)

        for changes, keys in change_groups.items():
            label = format_key_range(keys)
            for change in changes:
                sheet_lines.append(f"    [{label}] {change}")

        # Added rows
        added = [k for k in new_rows if k not in old_keys]
        if added:
            codes = [new_rows[k].get("Code", k) or k for k in added]
            if len(added) <= 3:
                for key in added:
                    code = new_rows[key].get("Code", key) or key
                    sheet_lines.append(f"    [{code}] ADDED: {format_row_short(new_rows[key], all_headers)}")
            else:
                sheet_lines.append(f"    {len(added)} rows added: {', '.join(codes)}")

        # Removed rows
        removed = [k for k in old_rows if k not in new_keys]
        if removed:
            codes = [old_rows[k].get("Code", k) or k for k in removed]
            if len(removed) <= 3:
                for key in removed:
                    code = old_rows[key].get("Code", key) or key
                    sheet_lines.append(f"    [{code}] REMOVED: {format_row_short(old_rows[key], all_headers)}")
            else:
                sheet_lines.append(f"    {len(removed)} rows removed: {', '.join(codes)}")

        if sheet_lines:
            has_changes = True
            print(f"  {sheet}")
            for line in sheet_lines:
                print(line)

    return has_changes


def main():
    project_dir = get_project_dir()
    os.chdir(project_dir)

    args = [a for a in sys.argv[1:] if a not in ("--set-path", "--show-path")]
    ref = DEFAULT_REF

    if len(args) >= 2 and args[0] == "--ref":
        ref = args[1]
        args = args[2:]

    if len(args) == 1:
        filepath = args[0]
        file_exists = os.path.isfile(filepath)
        tmp_path = get_ref_version(filepath, ref)

        if tmp_path is None and not file_exists:
            print(f"  (file not found locally or in {ref})")
            return

        if tmp_path is None:
            new_data = parse_xlsx(filepath)
            print(f"  (new file)")
            for sheet in sorted(new_data.keys()):
                rows = new_data[sheet]["rows"]
                if rows:
                    codes = [r.get("Code", "") or r.get("code", "") for r in rows.values()]
                    codes = [c for c in codes if c]
                    print(f"  {sheet} — {len(rows)} rows: {', '.join(codes)}")
            return

        if not file_exists:
            try:
                old_data = parse_xlsx(tmp_path)
                print(f"  (DELETED)")
                for sheet in sorted(old_data.keys()):
                    rows = old_data[sheet]["rows"]
                    if rows:
                        codes = [r.get("Code", "") or r.get("code", "") for r in rows.values()]
                        codes = [c for c in codes if c]
                        print(f"  {sheet} — had {len(rows)} rows: {', '.join(codes)}")
            finally:
                os.unlink(tmp_path)
            return

        try:
            old_data = parse_xlsx(tmp_path)
            new_data = parse_xlsx(filepath)
            if not diff_sheets(old_data, new_data):
                print("  (no differences)")
        finally:
            os.unlink(tmp_path)

    elif len(args) == 2:
        old_data = parse_xlsx(args[0])
        new_data = parse_xlsx(args[1])
        if not diff_sheets(old_data, new_data):
            print("  (no differences)")
    else:
        print(f"Usage: {sys.argv[0]} [--ref <git-ref>] <file.xlsx>", file=sys.stderr)
        print(f"       {sys.argv[0]} <old.xlsx> <new.xlsx>", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
