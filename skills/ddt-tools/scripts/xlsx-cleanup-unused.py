#!/usr/bin/env python3
"""Report unused codes in DDT reference data files.

Usage:
    xlsx-cleanup-unused.py              # report unused codes across all Data files

Latest version: https://github.com/Kimputing/centertest-skills/blob/main/skills/ddt-tools/scripts/xlsx-cleanup-unused.py
"""

import glob
import json
import os
import sys
from collections import defaultdict
from datetime import datetime

# Add script directory to path for config import
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ddt_config import get_project_dir

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def format_value(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)


def read_codes_from_sheet(wb, sheet_name):
    """Read all Code values from a sheet. Returns {code: row_index}."""
    ws = wb[sheet_name]
    headers = None
    codes = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [format_value(c) for c in row]
        if all(v == "" for v in values):
            continue
        if headers is None:
            headers = [v.lower() for v in values]
            continue
        row_dict = dict(zip(headers, values))
        code = row_dict.get("code", "")
        if code:
            codes[code] = row_idx
    return codes


def load_hierarchy():
    path = "testdata/DataDrivenHierarchy.json"
    if os.path.isfile(path):
        with open(path) as f:
            return json.load(f)
    return {}


def collect_all_references():
    """Scan all DC files and collect every referenced code per sheet name."""
    referenced = defaultdict(set)

    dc_files = glob.glob("testdata/**/*DC.xlsx", recursive=True) + glob.glob("testdata/*DC.xlsx")
    dc_files = [f for f in dc_files if not os.path.basename(f).startswith("~$")]
    dc_files = sorted(set(dc_files))

    for dc_path in dc_files:
        try:
            wb = load_workbook(dc_path, read_only=True, data_only=True)
        except Exception:
            continue

        dc_sheet = None
        for name in wb.sheetnames:
            if name.lower() == "datacombination":
                dc_sheet = name
                break

        if dc_sheet is None:
            wb.close()
            continue

        ws = wb[dc_sheet]
        headers = None
        for row in ws.iter_rows(values_only=True):
            values = [format_value(c) for c in row]
            if all(v == "" for v in values):
                continue
            if headers is None:
                headers = values
                continue
            row_dict = dict(zip(headers, values))
            for col in headers:
                if not col.startswith("#"):
                    continue
                cell_value = row_dict.get(col, "")
                if not cell_value:
                    continue
                sheet_name = col[1:].lower()
                for code in cell_value.split(","):
                    code = code.strip()
                    if code:
                        referenced[sheet_name].add(code)

        wb.close()

    return referenced


def collect_all_data_codes():
    """Scan all Data files and collect every code per file/sheet."""
    data_files = glob.glob("testdata/*Data.xlsx") + glob.glob("testdata/*Data.xlsx")
    data_files = sorted(set(f for f in data_files if not os.path.basename(f).startswith("~$")))

    result = {}
    for path in data_files:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        file_data = {}
        for sheet_name in wb.sheetnames:
            codes = read_codes_from_sheet(wb, sheet_name)
            if codes:
                file_data[sheet_name] = codes
        wb.close()
        result[path] = file_data

    return result


def main():
    project_dir = get_project_dir()
    os.chdir(project_dir)

    print("Scanning all DC files for references...")
    referenced = collect_all_references()

    print("Scanning all Data files for codes...")
    data_codes = collect_all_data_codes()

    total_unused = 0
    findings = []

    for filepath, sheets in sorted(data_codes.items()):
        for sheet_name, codes in sorted(sheets.items()):
            sheet_lower = sheet_name.lower()
            ref_codes = referenced.get(sheet_lower, set())
            unused = set(codes.keys()) - ref_codes
            if unused:
                findings.append((filepath, sheet_name, unused, codes))
                total_unused += len(unused)

    if not findings:
        print("\nNo unused codes found.")
        return

    print(f"\nFound {total_unused} unused code(s):\n")
    for filepath, sheet_name, unused, codes in findings:
        used_count = len(codes) - len(unused)
        basename = os.path.basename(filepath)
        if len(unused) == len(codes):
            tag = "ENTIRE SHEET UNUSED"
        else:
            tag = f"{used_count} used, {len(unused)} unused"
        print(f"  {basename} / {sheet_name} ({tag})")
        for code in sorted(unused):
            print(f"    - {code}")


if __name__ == "__main__":
    main()
