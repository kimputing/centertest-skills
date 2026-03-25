#!/usr/bin/env python3
"""Analyze CenterTest Data-Driven Testing structure and generate an Excel report.

Scans testdata/ for DC files, maps relationships between DataCombination files
and their referenced Data files, tracks code usage, and maps test classes to
datasources.

Usage:
    ddt-analyzer.py                         # analyze and generate report
    ddt-analyzer.py --exclude path1,path2   # exclude specific paths

Output: DDT_Analysis_<timestamp>.xlsx in the project's results/ directory.

Report sheets (15 total):
  Original:
   1. DataCombination_References  — which Data files each DC file references
   2. ReferencedFiles_DataCombination — which DC files reference each Data file
   3. Codes_Usage — aggregated code usage counts across all DC files
   4. Codes_Usage_Detail — per-DC-file code usage breakdown
   5. DataCombination_Tests — which test classes use each DC file as datasource
  New:
   6. Orphaned_DataFiles — xlsx files in testdata/ not referenced by any DC
   7. Broken_Datasources — @DataDriven annotations pointing to non-existent files
   8. Untested_DC_Files — DC files with no @DataDriven test method
   9. Unused_Codes — codes in Data files never referenced from any DC
  10. Hardcoded_DDTHelper — validation of DDTHelper.getXxx("literal") calls
  11. Hierarchy_Validation — DataDrivenHierarchy.json integrity checks
  12. Code_Coverage — % of codes used per Data file sheet
  13. Duplicate_Codes — same code appearing in multiple Data files/sheets
  14. DC_Metrics — complexity metrics per DC file
  15. Impact_Analysis — blast radius of each Data file

Latest version: https://github.com/Kimputing/centertest-skills/blob/main/skills/ddt-analyzer/scripts/ddt-analyzer.py
"""

import glob
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime

# Add script directory to path for shared config
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "ddt-tools", "scripts"))

try:
    from ddt_config import get_project_dir
except ImportError:
    def get_project_dir():
        path = os.environ.get("CENTERTEST_PROJECT_DIR")
        if path:
            return path
        config_file = os.path.expanduser("~/.centertest/ddt-tools.json")
        if os.path.isfile(config_file):
            with open(config_file) as f:
                config = json.load(f)
            if config.get("project_dir"):
                return config["project_dir"]
        print("No CenterTest project path configured.")
        print("Enter the path to the CenterTest project root:")
        path = input("> ").strip()
        if not path or not os.path.isdir(os.path.expanduser(path)):
            print("Error: invalid path")
            sys.exit(1)
        return os.path.expanduser(path)

try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# =============================================================================
# Data model
# =============================================================================

def format_value(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)


class SheetData:
    def __init__(self, name, codes=None):
        self.name = name
        self.codes = codes or set()


class ReferencedFile:
    def __init__(self, path):
        self.path = path
        self.sheets = []  # list of SheetData


class DataCombinationFile:
    def __init__(self, path):
        self.path = path
        self.referenced_files = []  # list of ReferencedFile
        self.references = {}  # {sheet_name: [code_values_from_dc_rows]}
        self.dc_codes = []  # Code column values from DataCombination sheet
        self.ref_column_count = 0  # number of # columns


# =============================================================================
# Parsing
# =============================================================================

_ref_file_cache = {}


def find_all_xlsx_files():
    """Find all xlsx files in testdata/."""
    files = glob.glob("testdata/**/*.xlsx", recursive=True) + glob.glob("testdata/*.xlsx")
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    return sorted(set(files))


def find_dc_files(exclude_paths=None):
    """Find all xlsx files with datacombination and references sheets."""
    exclude_paths = exclude_paths or []
    files = find_all_xlsx_files()

    dc_files = []
    for f in files:
        if any(f.startswith(ep) for ep in exclude_paths):
            continue
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
            sheet_names_lower = [s.lower() for s in wb.sheetnames]
            has_dc = "datacombination" in sheet_names_lower
            has_refs = "references" in sheet_names_lower
            wb.close()
            if has_dc and has_refs:
                dc_files.append(f)
        except Exception:
            continue

    return dc_files


def parse_dc_file(dc_path):
    """Parse a DataCombination file and return a DataCombinationFile object."""
    dc = DataCombinationFile(dc_path)

    wb = load_workbook(dc_path, read_only=True, data_only=True)

    dc_sheet_name = None
    ref_sheet_name = None
    for name in wb.sheetnames:
        if name.lower() == "datacombination":
            dc_sheet_name = name
        if name.lower() == "references":
            ref_sheet_name = name

    if dc_sheet_name:
        ws = wb[dc_sheet_name]
        headers = None
        ref_columns = {}
        code_col_idx = None
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            values = [format_value(c) for c in row]
            if all(v == "" for v in values):
                continue
            if headers is None:
                headers = values
                for i, h in enumerate(headers):
                    if h.startswith("#"):
                        ref_columns[i] = h[1:]
                        dc.references[h[1:]] = []
                    if h.lower() == "code":
                        code_col_idx = i
                dc.ref_column_count = len(ref_columns)
                continue
            # Collect Code column value
            if code_col_idx is not None and code_col_idx < len(values):
                code_val = values[code_col_idx]
                if code_val:
                    dc.dc_codes.append(code_val)
            for col_idx, sheet_name in ref_columns.items():
                val = values[col_idx] if col_idx < len(values) else ""
                if val:
                    for code in val.split(","):
                        code = code.strip()
                        if code:
                            dc.references[sheet_name].append(code)

    if ref_sheet_name:
        ws = wb[ref_sheet_name]
        headers = None
        for row in ws.iter_rows(values_only=True):
            values = [format_value(c) for c in row]
            if all(v == "" for v in values):
                continue
            if headers is None:
                headers = [v.lower().strip() for v in values]
                continue
            row_dict = dict(zip(headers, values))
            location = row_dict.get("location", "")
            if location:
                ref_path = os.path.normpath(os.path.join(os.path.dirname(dc_path), location))
                if not os.path.isfile(ref_path):
                    ref_path = location
                if os.path.isfile(ref_path):
                    ref_file = parse_referenced_file(ref_path)
                    dc.referenced_files.append(ref_file)

    wb.close()
    return dc


def parse_referenced_file(path):
    """Parse a referenced data file and extract codes per sheet."""
    if path in _ref_file_cache:
        return _ref_file_cache[path]

    ref = ReferencedFile(path)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = None
            codes = set()
            for row in ws.iter_rows(values_only=True):
                values = [format_value(c) for c in row]
                if all(v == "" for v in values):
                    continue
                if headers is None:
                    headers = [v.lower() for v in values]
                    continue
                row_dict = dict(zip(headers, values))
                code = row_dict.get("code", "")
                if code:
                    codes.add(code)
            if codes:
                ref.sheets.append(SheetData(sheet_name, codes))
        wb.close()
    except Exception:
        pass

    _ref_file_cache[path] = ref
    return ref


def find_test_datasource_mappings():
    """Scan Java source for @DataDriven(datasource = "...") annotations."""
    test_map = defaultdict(list)

    java_files = glob.glob("src/**/*.java", recursive=True)
    dd_pattern = re.compile(r'@DataDriven\s*\(\s*datasource\s*=\s*"([^"]+)"')
    class_pattern = re.compile(r'(?:public\s+)?(?:final\s+)?class\s+(\w+)')
    method_pattern = re.compile(r'(?:public|protected|private)?\s*(?:void|[A-Za-z]\w*)\s+(\w+)\s*\(')

    for java_file in java_files:
        try:
            with open(java_file) as f:
                content = f.read()
        except Exception:
            continue

        class_match = class_pattern.search(content)
        if not class_match:
            continue
        class_name = class_match.group(1)

        pkg_match = re.search(r'package\s+([\w.]+)\s*;', content)
        fqcn = f"{pkg_match.group(1)}.{class_name}" if pkg_match else class_name

        for dd_match in dd_pattern.finditer(content):
            datasource = dd_match.group(1)
            remaining = content[dd_match.end():]
            method_match = method_pattern.search(remaining)
            if method_match:
                method_name = method_match.group(1)
                test_map[datasource].append(f"{fqcn}#{method_name}")

    return dict(test_map)


def parse_ddthelper():
    """Parse DDTHelper.java to extract method -> (xlsx_file, sheet_or_none) mapping."""
    candidates = glob.glob("src/**/DDTHelper.java", recursive=True)
    if not candidates:
        return {}

    method_map = {}
    with open(candidates[0]) as f:
        content = f.read()

    # DC methods
    dc_pattern = re.compile(
        r'public static \S+ (\w+)\(String code\).*?'
        r'getExcelInputFileByCode\("([^"]+)",\s*code\)'
    )
    for m in dc_pattern.finditer(content):
        method_map[m.group(1)] = (m.group(2), None)

    # Reference methods
    ref_pattern = re.compile(
        r'public static \S+ (\w+)\(String code\).*?'
        r'provideDataRecordAsExcelInput\("([^"]+)",\s*"([^"]+)",\s*code\)'
    )
    for m in ref_pattern.finditer(content):
        method_map[m.group(1)] = (m.group(2), m.group(3))

    return method_map


def find_ddthelper_calls():
    """Find all DDTHelper.method("literal") calls in Java source."""
    pattern = re.compile(r'DDTHelper\.(\w+)\(\s*"([^"]+)"\s*\)')
    calls = []

    for java_file in glob.glob("src/**/*.java", recursive=True):
        if "DDTHelper.java" in java_file:
            continue
        try:
            with open(java_file) as f:
                for line_num, line in enumerate(f, start=1):
                    for m in pattern.finditer(line):
                        calls.append((java_file, line_num, m.group(1), m.group(2)))
        except Exception:
            continue

    return calls


def load_codes_from_xlsx(xlsx_path, sheet_name=None):
    """Load codes from an xlsx file. Returns dict {lowercase: original} or None."""
    if not os.path.isfile(xlsx_path):
        return None
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return None

    target = None
    if sheet_name is None:
        for name in wb.sheetnames:
            if name.lower() == "datacombination":
                target = name
                break
    else:
        for name in wb.sheetnames:
            if name.lower() == sheet_name.lower():
                target = name
                break

    if target is None:
        wb.close()
        return {}

    ws = wb[target]
    headers = None
    codes = {}
    for row in ws.iter_rows(values_only=True):
        values = [format_value(c) for c in row]
        if all(v == "" for v in values):
            continue
        if headers is None:
            headers = [v.lower() for v in values]
            continue
        row_dict = dict(zip(headers, values))
        code = row_dict.get("code", "")
        if code:
            codes[code.lower()] = code

    wb.close()
    return codes


# =============================================================================
# New analysis functions
# =============================================================================

def analyze_orphaned_files(dc_files_data):
    """Find xlsx files in testdata/ not referenced by any DC."""
    all_files = set(find_all_xlsx_files())
    dc_paths = set(dc.path for dc in dc_files_data)
    referenced_paths = set()
    for dc in dc_files_data:
        for ref in dc.referenced_files:
            referenced_paths.add(ref.path)

    known = dc_paths | referenced_paths
    orphaned = sorted(all_files - known)
    return orphaned


def analyze_broken_datasources(test_map):
    """Find @DataDriven datasources pointing to non-existent files."""
    broken = []
    for datasource, tests in test_map.items():
        if not os.path.isfile(datasource):
            broken.append((datasource, tests))
    return broken


def analyze_untested_dc_files(dc_files_data, test_map):
    """Find DC files that have no @DataDriven test referencing them."""
    tested_datasources = set(test_map.keys())
    untested = []
    for dc in dc_files_data:
        if dc.path not in tested_datasources:
            untested.append(dc.path)
    return sorted(untested)


def analyze_unused_codes(dc_files_data):
    """Find codes in Data files never referenced from any DC."""
    # Collect all referenced codes per sheet name (case-insensitive)
    referenced = defaultdict(set)  # sheet_name_lower -> set of code_lower
    for dc in dc_files_data:
        for sheet_name, codes in dc.references.items():
            for code in codes:
                referenced[sheet_name.lower()].add(code.lower())

    unused = []  # (file_path, sheet_name, code, total_codes, used_count)
    for ref_path, ref_file in sorted(_ref_file_cache.items()):
        for sheet in ref_file.sheets:
            ref_codes = referenced.get(sheet.name.lower(), set())
            sheet_unused = [c for c in sorted(sheet.codes) if c.lower() not in ref_codes]
            sheet_used = len(sheet.codes) - len(sheet_unused)
            for code in sheet_unused:
                unused.append((ref_path, sheet.name, code, len(sheet.codes), sheet_used))

    return unused


def analyze_hardcoded_ddthelper(ddthelper_map, ddthelper_calls):
    """Validate hardcoded DDTHelper calls against xlsx files."""
    codes_cache = {}
    results = []  # (file, line, method, code, status, available)

    for java_file, line_num, method_name, code_literal in ddthelper_calls:
        if method_name not in ddthelper_map:
            results.append((java_file, line_num, method_name, code_literal, "UNKNOWN_METHOD", ""))
            continue

        xlsx_file, sheet_name = ddthelper_map[method_name]
        cache_key = (xlsx_file, sheet_name)

        if cache_key not in codes_cache:
            codes_cache[cache_key] = load_codes_from_xlsx(xlsx_file, sheet_name)

        available = codes_cache[cache_key]

        if available is None:
            results.append((java_file, line_num, method_name, code_literal, "FILE_NOT_FOUND", xlsx_file))
            continue

        if code_literal.lower() in available:
            results.append((java_file, line_num, method_name, code_literal, "VALID", ""))
        else:
            avail_str = ", ".join(sorted(available.values()))
            results.append((java_file, line_num, method_name, code_literal, "INVALID", avail_str))

    return results


def analyze_hierarchy():
    """Validate DataDrivenHierarchy.json."""
    hierarchy_path = "testdata/DataDrivenHierarchy.json"
    results = []  # (item, status, detail)

    if not os.path.isfile(hierarchy_path):
        results.append(("DataDrivenHierarchy.json", "MISSING", "File not found"))
        return results

    with open(hierarchy_path) as f:
        hierarchy = json.load(f)

    all_xlsx = set(find_all_xlsx_files())

    # Check each parent and its child patterns
    for parent_dc, child_patterns in hierarchy.items():
        # Check parent exists
        if not os.path.isfile(parent_dc):
            results.append((parent_dc, "PARENT_MISSING", "Parent DC file not found"))
        else:
            results.append((parent_dc, "OK", f"Parent exists, {len(child_patterns)} pattern(s)"))

        # Check which files each pattern matches
        for pattern in child_patterns:
            matched = [f for f in all_xlsx if re.match(pattern, f)]
            if not matched:
                results.append((f"  pattern: {pattern}", "NO_MATCH", "No files match this pattern"))
            else:
                for m in matched:
                    results.append((f"  {m}", "MATCHED", f"Matched by {pattern}"))

    # Check for child DC files NOT covered by any hierarchy pattern
    all_patterns = []
    for patterns in hierarchy.values():
        all_patterns.extend(patterns)

    dc_files = find_dc_files()
    parent_dcs = set(hierarchy.keys())
    for dc in dc_files:
        if dc in parent_dcs:
            continue
        covered = any(re.match(p, dc) for p in all_patterns)
        if not covered:
            results.append((dc, "UNCOVERED", "DC file not covered by any hierarchy pattern"))

    return results


def analyze_code_coverage(dc_files_data, ddthelper_map, ddthelper_calls):
    """Calculate code coverage per Data file sheet."""
    # Codes used by DC files
    dc_used = defaultdict(set)  # (ref_path, sheet_name_lower) -> set of code_lower
    for dc in dc_files_data:
        for ref in dc.referenced_files:
            for sheet_name, codes in dc.references.items():
                for code in codes:
                    dc_used[(ref.path, sheet_name.lower())].add(code.lower())

    # Codes used by hardcoded DDTHelper
    hc_used = defaultdict(set)  # (xlsx_file, sheet_name_lower) -> set of code_lower
    for _, _, method_name, code_literal in ddthelper_calls:
        if method_name in ddthelper_map:
            xlsx_file, sheet_name = ddthelper_map[method_name]
            sheet_key = (sheet_name or "datacombination").lower()
            hc_used[(xlsx_file, sheet_key)].add(code_literal.lower())

    results = []  # (file, sheet, total, dc_used, hc_used, unused, pct)
    for ref_path, ref_file in sorted(_ref_file_cache.items()):
        for sheet in ref_file.sheets:
            total = len(sheet.codes)
            dc_count = len(dc_used.get((ref_path, sheet.name.lower()), set()))
            hc_count = len(hc_used.get((ref_path, sheet.name.lower()), set()))
            all_used = dc_used.get((ref_path, sheet.name.lower()), set()) | hc_used.get((ref_path, sheet.name.lower()), set())
            actual_used = len([c for c in sheet.codes if c.lower() in all_used])
            unused = total - actual_used
            pct = round(actual_used / total * 100, 1) if total > 0 else 0
            results.append((ref_path, sheet.name, total, dc_count, hc_count, unused, pct))

    return results


def analyze_duplicate_codes():
    """Find codes appearing in multiple Data files/sheets."""
    # code_lower -> [(file, sheet, original_code)]
    code_locations = defaultdict(list)
    for ref_path, ref_file in _ref_file_cache.items():
        for sheet in ref_file.sheets:
            for code in sheet.codes:
                code_locations[code.lower()].append((ref_path, sheet.name, code))

    duplicates = []
    for code_lower, locations in sorted(code_locations.items()):
        if len(locations) > 1:
            for file_path, sheet_name, original in locations:
                duplicates.append((original, file_path, sheet_name, len(locations)))

    return duplicates


def analyze_dc_metrics(dc_files_data, test_map, ddthelper_map, ddthelper_calls):
    """Calculate complexity metrics per DC file."""
    # Count hardcoded calls per DC file
    hc_per_dc = defaultdict(int)  # xlsx_path -> count
    for _, _, method_name, _ in ddthelper_calls:
        if method_name in ddthelper_map:
            xlsx_file, _ = ddthelper_map[method_name]
            hc_per_dc[xlsx_file] += 1

    results = []  # (path, dc_codes, ref_columns, ref_files, test_count, hc_count)
    for dc in dc_files_data:
        test_count = len(test_map.get(dc.path, []))
        hc_count = hc_per_dc.get(dc.path, 0)
        total_code_usages = sum(len(codes) for codes in dc.references.values())
        results.append((
            dc.path,
            len(dc.dc_codes),
            dc.ref_column_count,
            len(dc.referenced_files),
            total_code_usages,
            test_count,
            hc_count,
        ))

    # Sort by total complexity (codes * refs * tests)
    results.sort(key=lambda r: r[1] * r[3] * max(r[5], 1), reverse=True)
    return results


def analyze_impact(dc_files_data, test_map, ddthelper_map, ddthelper_calls):
    """Calculate blast radius for each Data file."""
    # Build reverse maps
    ref_to_dcs = defaultdict(set)
    for dc in dc_files_data:
        for ref in dc.referenced_files:
            ref_to_dcs[ref.path].add(dc.path)

    # Tests per Data file (transitive through DCs)
    ref_to_tests = defaultdict(set)
    for ref_path, dc_paths in ref_to_dcs.items():
        for dc_path in dc_paths:
            for test in test_map.get(dc_path, []):
                ref_to_tests[ref_path].add(test)

    # Hardcoded calls per Data file
    ref_to_hc = defaultdict(list)
    for java_file, line_num, method_name, code_literal in ddthelper_calls:
        if method_name in ddthelper_map:
            xlsx_file, _ = ddthelper_map[method_name]
            ref_to_hc[xlsx_file].append(f"{java_file}:{line_num}")

    results = []  # (file, dc_count, test_count, hc_count, dc_list, test_list, hc_list)
    for ref_path in sorted(_ref_file_cache.keys()):
        dcs = sorted(ref_to_dcs.get(ref_path, set()))
        tests = sorted(ref_to_tests.get(ref_path, set()))
        hcs = sorted(ref_to_hc.get(ref_path, []))
        results.append((ref_path, len(dcs), len(tests), len(hcs),
                         "; ".join(dcs), "; ".join(tests), "; ".join(hcs)))

    results.sort(key=lambda r: r[1] + r[2] + r[3], reverse=True)
    return results


# =============================================================================
# Report generation
# =============================================================================

def generate_report(dc_files_data, test_map, ddthelper_map, ddthelper_calls):
    """Generate the Excel analysis report with all 15 sheets."""
    wb = Workbook()

    # --- Sheet 1: DataCombination_References ---
    ws = wb.active
    ws.title = "DC_References"

    all_ref_files = set()
    for dc in dc_files_data:
        for ref in dc.referenced_files:
            all_ref_files.add(ref.path)
    all_ref_files = sorted(all_ref_files)

    ws.append(["Data Combination File"] + [os.path.basename(r) for r in all_ref_files])
    ref_index = {r: i + 1 for i, r in enumerate(all_ref_files)}
    for dc in dc_files_data:
        row = [""] * (len(all_ref_files) + 1)
        row[0] = dc.path
        for ref in dc.referenced_files:
            idx = ref_index.get(ref.path)
            if idx:
                row[idx] = "X"
        ws.append(row)

    # --- Sheet 2: ReferencedFiles_DataCombination ---
    ws = wb.create_sheet("RefFiles_DC")
    ref_to_dc = defaultdict(list)
    for dc in dc_files_data:
        for ref in dc.referenced_files:
            ref_to_dc[ref.path].append(dc.path)

    all_dc_paths = sorted(set(dc.path for dc in dc_files_data))
    ws.append(["Referenced File"] + [os.path.basename(d) for d in all_dc_paths])
    dc_index = {d: i + 1 for i, d in enumerate(all_dc_paths)}
    for ref_path in sorted(ref_to_dc.keys()):
        row = [""] * (len(all_dc_paths) + 1)
        row[0] = ref_path
        for dc_path in ref_to_dc[ref_path]:
            idx = dc_index.get(dc_path)
            if idx:
                row[idx] = "X"
        ws.append(row)

    # --- Sheet 3: Codes_Usage ---
    ws = wb.create_sheet("Codes_Usage")
    ws.append(["Referenced File", "Sheet", "Code", "Count"])
    for ref_path, ref_file in sorted(_ref_file_cache.items()):
        for sheet in ref_file.sheets:
            for code in sorted(sheet.codes):
                count = 0
                for dc in dc_files_data:
                    if any(r.path == ref_path for r in dc.referenced_files):
                        code_list = dc.references.get(sheet.name, [])
                        count += sum(1 for c in code_list if c.lower() == code.lower())
                ws.append([ref_path, sheet.name, code, count])

    # --- Sheet 4: Codes_Usage_Detail ---
    ws = wb.create_sheet("Codes_Usage_Detail")
    ws.append(["Referenced File", "Sheet", "Code", "Data Combination File"])
    for ref_path, ref_file in sorted(_ref_file_cache.items()):
        for sheet in ref_file.sheets:
            for code in sorted(sheet.codes):
                for dc in dc_files_data:
                    if any(r.path == ref_path for r in dc.referenced_files):
                        code_list = dc.references.get(sheet.name, [])
                        if any(c.lower() == code.lower() for c in code_list):
                            ws.append([ref_path, sheet.name, code, dc.path])

    # --- Sheet 5: DataCombination_Tests ---
    ws = wb.create_sheet("DC_Tests")
    ws.append(["Data Combination File", "Test"])
    for datasource, tests in sorted(test_map.items()):
        for test in sorted(tests):
            ws.append([datasource, test])

    # --- Sheet 6: Orphaned_DataFiles ---
    ws = wb.create_sheet("Orphaned_DataFiles")
    ws.append(["File", "Reason"])
    orphaned = analyze_orphaned_files(dc_files_data)
    for f in orphaned:
        ws.append([f, "Not referenced by any DC file"])

    # --- Sheet 7: Broken_Datasources ---
    ws = wb.create_sheet("Broken_Datasources")
    ws.append(["Datasource Path", "Test Methods"])
    broken = analyze_broken_datasources(test_map)
    for datasource, tests in broken:
        ws.append([datasource, "; ".join(tests)])

    # --- Sheet 8: Untested_DC_Files ---
    ws = wb.create_sheet("Untested_DC_Files")
    ws.append(["DC File", "Reason"])
    untested = analyze_untested_dc_files(dc_files_data, test_map)
    for f in untested:
        ws.append([f, "No @DataDriven test references this file"])

    # --- Sheet 9: Unused_Codes ---
    ws = wb.create_sheet("Unused_Codes")
    ws.append(["Data File", "Sheet", "Unused Code", "Total Codes", "Used Codes"])
    unused = analyze_unused_codes(dc_files_data)
    for file_path, sheet_name, code, total, used in unused:
        ws.append([file_path, sheet_name, code, total, used])

    # --- Sheet 10: Hardcoded_DDTHelper ---
    ws = wb.create_sheet("Hardcoded_DDTHelper")
    ws.append(["Java File", "Line", "Method", "Code", "Status", "Available Codes"])
    hc_results = analyze_hardcoded_ddthelper(ddthelper_map, ddthelper_calls)
    for java_file, line_num, method, code, status, available in hc_results:
        ws.append([java_file, line_num, method, code, status, available])

    # --- Sheet 11: Hierarchy_Validation ---
    ws = wb.create_sheet("Hierarchy_Validation")
    ws.append(["Item", "Status", "Detail"])
    hierarchy_results = analyze_hierarchy()
    for item, status, detail in hierarchy_results:
        ws.append([item, status, detail])

    # --- Sheet 12: Code_Coverage ---
    ws = wb.create_sheet("Code_Coverage")
    ws.append(["Data File", "Sheet", "Total Codes", "Used by DC", "Used by Hardcoded", "Unused", "Coverage %"])
    coverage = analyze_code_coverage(dc_files_data, ddthelper_map, ddthelper_calls)
    for row in coverage:
        ws.append(list(row))

    # --- Sheet 13: Duplicate_Codes ---
    ws = wb.create_sheet("Duplicate_Codes")
    ws.append(["Code", "Data File", "Sheet", "Occurrences"])
    duplicates = analyze_duplicate_codes()
    for code, file_path, sheet_name, count in duplicates:
        ws.append([code, file_path, sheet_name, count])

    # --- Sheet 14: DC_Metrics ---
    ws = wb.create_sheet("DC_Metrics")
    ws.append(["DC File", "DC Codes", "# Columns", "Ref Files", "Code Usages", "Tests", "Hardcoded Calls"])
    metrics = analyze_dc_metrics(dc_files_data, test_map, ddthelper_map, ddthelper_calls)
    for row in metrics:
        ws.append(list(row))

    # --- Sheet 15: Impact_Analysis ---
    ws = wb.create_sheet("Impact_Analysis")
    ws.append(["Data File", "DC Files", "Tests", "Hardcoded Calls", "DC File List", "Test List", "Hardcoded Locations"])
    impact = analyze_impact(dc_files_data, test_map, ddthelper_map, ddthelper_calls)
    for row in impact:
        ws.append(list(row))

    # Save
    os.makedirs("results", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"results/DDT_Analysis_{timestamp}.xlsx"
    wb.save(output_path)
    return output_path


# =============================================================================
# Main
# =============================================================================

def main():
    project_dir = get_project_dir()
    os.chdir(project_dir)

    exclude_paths = []
    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if args[i] == "--exclude" and i + 1 < len(args):
            exclude_paths = [p.strip() for p in args[i + 1].split(",")]
            i += 2
        else:
            i += 1

    print("DDT Analyzer")
    print("=" * 60)

    # Step 1: Find DC files
    print("\n[1/6] Scanning for DataCombination files...")
    dc_paths = find_dc_files(exclude_paths)
    print(f"  Found {len(dc_paths)} DC files")

    # Step 2: Parse each DC file
    print("\n[2/6] Parsing DC files and referenced data...")
    dc_files_data = []
    for dc_path in dc_paths:
        try:
            dc = parse_dc_file(dc_path)
            dc_files_data.append(dc)
            ref_count = len(dc.referenced_files)
            code_count = sum(len(codes) for codes in dc.references.values())
            print(f"  {dc_path}: {len(dc.dc_codes)} codes, {ref_count} refs, {code_count} code usages")
        except Exception as e:
            print(f"  {dc_path}: ERROR — {e}")

    print(f"  Total referenced data files: {len(_ref_file_cache)}")

    # Step 3: Find test mappings
    print("\n[3/6] Scanning Java source for @DataDriven tests...")
    test_map = find_test_datasource_mappings()
    total_tests = sum(len(v) for v in test_map.values())
    print(f"  Found {total_tests} test methods across {len(test_map)} datasources")

    # Step 4: Parse DDTHelper
    print("\n[4/6] Parsing DDTHelper for hardcoded code validation...")
    ddthelper_map = parse_ddthelper()
    print(f"  Found {len(ddthelper_map)} DDTHelper methods")

    # Step 5: Find hardcoded calls
    print("\n[5/6] Scanning for hardcoded DDTHelper calls...")
    ddthelper_calls = find_ddthelper_calls()
    print(f"  Found {len(ddthelper_calls)} hardcoded DDTHelper.getXxx(\"code\") calls")

    # Step 6: Generate report
    print("\n[6/6] Running analysis and generating report...")

    orphaned = analyze_orphaned_files(dc_files_data)
    print(f"  Orphaned files:        {len(orphaned)}")

    broken = analyze_broken_datasources(test_map)
    print(f"  Broken datasources:    {len(broken)}")

    untested = analyze_untested_dc_files(dc_files_data, test_map)
    print(f"  Untested DC files:     {len(untested)}")

    unused = analyze_unused_codes(dc_files_data)
    print(f"  Unused codes:          {len(unused)}")

    hc_results = analyze_hardcoded_ddthelper(ddthelper_map, ddthelper_calls)
    invalid_hc = [r for r in hc_results if r[4] == "INVALID"]
    print(f"  Invalid hardcoded:     {len(invalid_hc)}")

    hierarchy = analyze_hierarchy()
    hierarchy_issues = [r for r in hierarchy if r[1] not in ("OK", "MATCHED")]
    print(f"  Hierarchy issues:      {len(hierarchy_issues)}")

    duplicates = analyze_duplicate_codes()
    unique_dup_codes = len(set(d[0].lower() for d in duplicates))
    print(f"  Duplicate codes:       {unique_dup_codes}")

    output_path = generate_report(dc_files_data, test_map, ddthelper_map, ddthelper_calls)

    # Summary
    print(f"\n{'=' * 60}")
    print("Summary:")
    print(f"  DC files:          {len(dc_files_data)}")
    print(f"  Referenced files:  {len(_ref_file_cache)}")
    print(f"  Test methods:      {total_tests}")
    print(f"  Report sheets:     15")
    print(f"  Report:            {output_path}")

    issues = len(orphaned) + len(broken) + len(invalid_hc) + len(hierarchy_issues)
    if issues > 0:
        print(f"\n  ⚠ {issues} issue(s) found — check report for details")


if __name__ == "__main__":
    main()
